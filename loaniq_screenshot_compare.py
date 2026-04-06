"""
LoanIQ Screenshot Comparison Tool
==================================
Compares screenshots from two LoanIQ versions (e.g. 7.7.1.0 vs 7.7.3.0)
organised in the same folder structure.

Uses EasyOCR (no external binary needed) to extract text/data from each
screenshot, compares the content, and writes anomalies to an Excel report.

Requirements (install once):
    pip3 install easyocr pillow openpyxl

Usage:
    python3 loaniq_screenshot_compare.py \\
        --v1 "/path/to/LoanIQ_7.7.1.0" \\
        --v2 "/path/to/LoanIQ_7.7.3.0" \\
        --out "comparison_report.xlsx" \\
        --threshold 95

    --threshold  (optional, default 95) minimum similarity % to consider
                 two screenshots "matching". Below this → flagged as anomaly.
    --verbose    print every file as it is processed
"""

import argparse
import difflib
import os
import re
import sys
import textwrap
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Dependency checks
# ---------------------------------------------------------------------------

try:
    from PIL import Image
except ImportError:
    sys.exit("❌  Pillow not found. Run:  pip3 install pillow")

try:
    import easyocr
except ImportError:
    sys.exit("❌  EasyOCR not found. Run:  pip3 install easyocr")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("❌  openpyxl not found. Run:  pip3 install openpyxl")


# ---------------------------------------------------------------------------
# Globals
# ---------------------------------------------------------------------------

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".tif", ".gif"}

# EasyOCR reader is initialised once and reused (model loads ~100 MB on first run)
_ocr_reader = None

def get_reader():
    global _ocr_reader
    if _ocr_reader is None:
        print("\n⏳  Loading EasyOCR model (first run downloads ~100 MB — one-time only)…")
        _ocr_reader = easyocr.Reader(["en"], gpu=False, verbose=False)
        print("    Model ready.\n")
    return _ocr_reader


# ---------------------------------------------------------------------------
# OCR helpers
# ---------------------------------------------------------------------------

def extract_text(image_path: Path) -> str:
    """Return all visible text from *image_path* using EasyOCR."""
    try:
        reader = get_reader()
        # detail=0 returns plain strings; paragraph=True merges nearby words into lines
        results = reader.readtext(str(image_path), detail=0, paragraph=True)
        return "\n".join(results).strip()
    except Exception as exc:
        return f"[OCR_ERROR: {exc}]"


def normalize_text(text: str) -> list:
    """Split OCR output into cleaned, non-empty lines for comparison."""
    lines = []
    for line in text.splitlines():
        line = re.sub(r"[ \t]+", " ", line).strip()
        if line:
            lines.append(line)
    return lines


def text_similarity(lines_a: list, lines_b: list) -> float:
    """Return a 0–100 similarity score between two line lists."""
    seq = difflib.SequenceMatcher(None, lines_a, lines_b)
    return round(seq.ratio() * 100, 2)


def build_diff(lines_a: list, lines_b: list):
    """Return (lines_only_in_a, lines_only_in_b) — content that changed."""
    matcher = difflib.SequenceMatcher(None, lines_a, lines_b)
    only_in_a, only_in_b = [], []
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag in ("delete", "replace"):
            only_in_a.extend(lines_a[i1:i2])
        if tag in ("insert", "replace"):
            only_in_b.extend(lines_b[j1:j2])
    return only_in_a, only_in_b


# ---------------------------------------------------------------------------
# Folder traversal
# ---------------------------------------------------------------------------

def collect_images(root: Path) -> dict:
    """Walk *root* recursively and return {relative_path: absolute_Path}."""
    images = {}
    for path in root.rglob("*"):
        if path.suffix.lower() in IMAGE_EXTENSIONS and path.is_file():
            rel = path.relative_to(root).as_posix()
            images[rel] = path
    return images


# ---------------------------------------------------------------------------
# Comparison engine
# ---------------------------------------------------------------------------

def compare_folders(v1_root: Path, v2_root: Path, threshold: float, verbose: bool) -> dict:
    print(f"\n📂  Scanning v1 folder: {v1_root}")
    v1_images = collect_images(v1_root)
    print(f"    Found {len(v1_images)} image(s).")

    print(f"📂  Scanning v2 folder: {v2_root}")
    v2_images = collect_images(v2_root)
    print(f"    Found {len(v2_images)} image(s).")

    all_keys    = sorted(set(v1_images) | set(v2_images))
    only_in_v1  = sorted(k for k in all_keys if k in v1_images and k not in v2_images)
    only_in_v2  = sorted(k for k in all_keys if k in v2_images and k not in v1_images)
    common      = sorted(k for k in all_keys if k in v1_images and k in v2_images)

    print(f"\n📊  Files in common : {len(common)}")
    print(f"    Only in v1      : {len(only_in_v1)}")
    print(f"    Only in v2      : {len(only_in_v2)}")
    print(f"\n🔍  Running OCR comparison on {len(common)} matched pair(s)…")

    # Initialise the reader before the loop so the load message prints once
    get_reader()

    matched_results = []
    for idx, rel in enumerate(common, 1):
        p1, p2 = v1_images[rel], v2_images[rel]

        if verbose:
            print(f"  [{idx:>4}/{len(common)}] {rel}")
        else:
            # Show a simple progress indicator
            print(f"  Processing {idx}/{len(common)}…", end="\r")

        text1  = extract_text(p1)
        text2  = extract_text(p2)
        lines1 = normalize_text(text1)
        lines2 = normalize_text(text2)

        similarity         = text_similarity(lines1, lines2)
        only_in_1, only_in_2 = build_diff(lines1, lines2)
        is_anomaly         = similarity < threshold

        if is_anomaly:
            print(f"  ⚠  {rel}  →  {similarity:.1f}% similarity")

        matched_results.append({
            "relative_path"  : rel,
            "v1_path"        : str(p1),
            "v2_path"        : str(p2),
            "similarity_pct" : similarity,
            "is_anomaly"     : is_anomaly,
            "lines_v1_total" : len(lines1),
            "lines_v2_total" : len(lines2),
            "text_removed"   : "\n".join(only_in_1),
            "text_added"     : "\n".join(only_in_2),
            "raw_text_v1"    : text1,
            "raw_text_v2"    : text2,
        })

    print()  # clear progress line
    anomaly_count = sum(1 for r in matched_results if r["is_anomaly"])
    total_issues  = anomaly_count + len(only_in_v1) + len(only_in_v2)
    print(f"✅  Comparison complete.  Total anomalies: {total_issues}")

    return {
        "matched_pairs" : matched_results,
        "only_in_v1"    : only_in_v1,
        "only_in_v2"    : only_in_v2,
    }


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

CLR_HEADER   = "1F3864"
CLR_HDR_FG   = "FFFFFF"
CLR_ANOMALY  = "FCE4D6"
CLR_OK       = "E2EFDA"
CLR_MISSING  = "FFF2CC"
CLR_TITLE_BG = "2E4057"

THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _hdr_font():
    return Font(name="Calibri", bold=True, color=CLR_HDR_FG, size=11)

def _cell_font(bold=False, color="000000", size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def _write_header_row(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = _hdr_font()
        c.fill      = _fill(CLR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BORDER


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def build_excel_report(results, v1_root, v2_root, output_path, threshold):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _sheet_summary(wb, results, v1_root, v2_root, threshold)
    _sheet_anomalies(wb, results)
    _sheet_missing(wb, results)
    _sheet_all(wb, results)

    wb.save(output_path)
    print(f"\n📄  Report saved → {output_path}")


def _sheet_summary(wb, results, v1_root, v2_root, threshold):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    matched       = results["matched_pairs"]
    anomalies     = [r for r in matched if r["is_anomaly"]]
    ok            = [r for r in matched if not r["is_anomaly"]]
    only_v1       = results["only_in_v1"]
    only_v2       = results["only_in_v2"]
    total_issues  = len(anomalies) + len(only_v1) + len(only_v2)

    # Title
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value     = "LoanIQ Screenshot Comparison Report"
    t.font      = Font(name="Calibri", bold=True, size=16, color=CLR_HDR_FG)
    t.fill      = _fill(CLR_TITLE_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:D2")
    s = ws["A2"]
    s.value     = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                   f"   |   Similarity threshold: {threshold}%")
    s.font      = Font(name="Calibri", italic=True, size=10, color="666666")
    s.alignment = Alignment(horizontal="center")

    kv_rows = [
        ("Version 1 folder",            str(v1_root)),
        ("Version 2 folder",            str(v2_root)),
        ("Matched pairs compared",       len(matched)),
        ("",                             ""),
        ("✓  Matching (no anomaly)",     len(ok)),
        ("⚠  Content anomalies",        len(anomalies)),
        ("⚠  Missing from v2",          len(only_v1)),
        ("⚠  New in v2 (not in v1)",    len(only_v2)),
        ("",                             ""),
        ("TOTAL ANOMALIES",              total_issues),
    ]

    for i, (key, val) in enumerate(kv_rows, 4):
        kc = ws.cell(row=i, column=1, value=key)
        vc = ws.cell(row=i, column=2, value=val)
        if key == "TOTAL ANOMALIES":
            color = "C00000" if total_issues else "375623"
            bg    = "FCE4D6" if total_issues else "E2EFDA"
            for c in [kc, vc]:
                c.font   = Font(name="Calibri", bold=True, size=12, color=color)
                c.fill   = _fill(bg)
                c.border = BORDER
        elif key.startswith("⚠"):
            kc.font = Font(name="Calibri", bold=True, size=10, color="C00000")
            vc.font = Font(name="Calibri", bold=True, size=10, color="C00000")
        elif key.startswith("✓"):
            kc.font = Font(name="Calibri", size=10, color="375623")
            vc.font = Font(name="Calibri", size=10, color="375623")
        else:
            kc.font = _cell_font(bold=True)
            vc.font = _cell_font()
        for c in [kc, vc]:
            c.alignment = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 80


def _sheet_anomalies(wb, results):
    ws = wb.create_sheet("⚠ Anomalies")
    ws.sheet_view.showGridLines = False

    headers = [
        "Relative Path", "Similarity %", "Lines in v1", "Lines in v2",
        "Text REMOVED (in v1, missing in v2)", "Text ADDED (new in v2)",
    ]
    _write_header_row(ws, headers)
    ws.row_dimensions[1].height = 30

    anomalies = [r for r in results["matched_pairs"] if r["is_anomaly"]]

    for row_idx, r in enumerate(anomalies, 2):
        bg = CLR_ANOMALY if row_idx % 2 == 0 else "FAD7C4"
        vals = [
            r["relative_path"],
            r["similarity_pct"],
            r["lines_v1_total"],
            r["lines_v2_total"],
            r["text_removed"][:2000] or "",
            r["text_added"][:2000]   or "",
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill      = _fill(bg)
            c.border    = BORDER
            c.font      = _cell_font(size=9)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.cell(row=row_idx, column=2).font = _cell_font(bold=True, color="C00000", size=9)
        ws.row_dimensions[row_idx].height = 60

    if not anomalies:
        ws.cell(row=2, column=1, value="🎉  No content anomalies detected!").font = \
            Font(name="Calibri", bold=True, color="375623", size=11)

    for col, w in enumerate([55, 14, 12, 12, 55, 55], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"


def _sheet_missing(wb, results):
    ws = wb.create_sheet("⚠ Missing Files")
    ws.sheet_view.showGridLines = False

    _write_header_row(ws, ["Relative Path", "In v1?", "In v2?", "Action Required"])
    ws.row_dimensions[1].height = 28

    rows = (
        [(p, "✓ Yes", "✗ Missing", "Screenshot removed or renamed in v2 — verify intentional")
         for p in results["only_in_v1"]] +
        [(p, "✗ Missing", "✓ Yes", "New screenshot in v2 — confirm this is expected")
         for p in results["only_in_v2"]]
    )

    for row_idx, (path, in_v1, in_v2, action) in enumerate(rows, 2):
        bg = CLR_MISSING if row_idx % 2 == 0 else "FEE9A0"
        for col, val in enumerate([path, in_v1, in_v2, action], 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill      = _fill(bg)
            c.border    = BORDER
            c.font      = _cell_font(size=9)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row_idx].height = 20

    if not rows:
        ws.cell(row=2, column=1, value="✓  All screenshots are present in both versions.").font = \
            Font(name="Calibri", bold=True, color="375623", size=11)

    for col, w in enumerate([70, 12, 12, 60], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"


def _sheet_all(wb, results):
    ws = wb.create_sheet("All Comparisons")
    ws.sheet_view.showGridLines = False

    headers = [
        "Status", "Relative Path", "Similarity %",
        "Lines v1", "Lines v2",
        "Text Removed (v1→v2)", "Text Added (v1→v2)",
    ]
    _write_header_row(ws, headers)
    ws.row_dimensions[1].height = 28

    all_rows = sorted(results["matched_pairs"],
                      key=lambda r: (not r["is_anomaly"], r["relative_path"]))

    for row_idx, r in enumerate(all_rows, 2):
        is_bad = r["is_anomaly"]
        bg     = CLR_ANOMALY if is_bad else (CLR_OK if row_idx % 2 == 0 else "D6EAD1")
        vals   = [
            "⚠ ANOMALY" if is_bad else "✓ OK",
            r["relative_path"],
            r["similarity_pct"],
            r["lines_v1_total"],
            r["lines_v2_total"],
            r["text_removed"][:1000] or "",
            r["text_added"][:1000]   or "",
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill      = _fill(bg)
            c.border    = BORDER
            c.font      = _cell_font(bold=is_bad, color="C00000" if is_bad else "000000", size=9)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row_idx].height = 45

    for col, w in enumerate([13, 55, 14, 10, 10, 55, 55], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description="Compare LoanIQ screenshots between two version folders using EasyOCR.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""\
            Example:
              python3 loaniq_screenshot_compare.py \\
                  --v1 "/Users/you/LoanIQ/7.7.1.0" \\
                  --v2 "/Users/you/LoanIQ/7.7.3.0" \\
                  --out "loaniq_comparison.xlsx" \\
                  --threshold 95
        """),
    )
    parser.add_argument("--v1",        required=True,  help="Root folder for LoanIQ v1 screenshots")
    parser.add_argument("--v2",        required=True,  help="Root folder for LoanIQ v2 screenshots")
    parser.add_argument("--out",       default="loaniq_comparison.xlsx",
                        help="Output Excel file (default: loaniq_comparison.xlsx)")
    parser.add_argument("--threshold", type=float, default=95.0,
                        help="Min similarity %% to pass (default: 95)")
    parser.add_argument("--verbose",   action="store_true",
                        help="Print every file as it is processed")
    return parser.parse_args()


def main():
    args = parse_args()

    v1_root     = Path(args.v1).expanduser().resolve()
    v2_root     = Path(args.v2).expanduser().resolve()
    output_path = Path(args.out).expanduser().resolve()

    if not v1_root.exists():
        sys.exit(f"❌  v1 folder not found: {v1_root}")
    if not v2_root.exists():
        sys.exit(f"❌  v2 folder not found: {v2_root}")

    print("=" * 65)
    print("  LoanIQ Screenshot Comparison Tool  (powered by EasyOCR)")
    print("=" * 65)
    print(f"  v1        : {v1_root}")
    print(f"  v2        : {v2_root}")
    print(f"  Threshold : {args.threshold}%")
    print(f"  Output    : {output_path}")
    print("=" * 65)

    results = compare_folders(v1_root, v2_root, args.threshold, args.verbose)
    build_excel_report(results, v1_root, v2_root, output_path, args.threshold)

    matched   = results["matched_pairs"]
    anomalies = [r for r in matched if r["is_anomaly"]]
    total     = len(anomalies) + len(results["only_in_v1"]) + len(results["only_in_v2"])

    print(f"\n{'='*65}")
    print(f"  RESULT: {total} anomaly(s) across {len(matched)} matched screenshot pair(s).")
    print(f"{'='*65}\n")


if __name__ == "__main__":
    main()
